import openpyxl
from twilio.rest import Client
from datetime import datetime, timedelta


# Inicialização da lista global de dados
lista_clientes = []
lista_nmr_clientes = []
lista_valores = []
lista_porcentagem = []
lista_pagamento = []
lista_parcelas = []
lista_valor_parcelas = []
lista_valor_total = []
lista_data_cadastro = []


# Função para coletar os dados do cliente
def coletar_dados_cliente():
    cliente = input("Digite o nome do cliente: ")
    lista_clientes.append(cliente)

    nmr_cliente = input("Digite o número do cliente (+55 + DDD + Cel): ")
    lista_nmr_clientes.append(nmr_cliente)

    valor = int(input("Digite o valor emprestado ao cliente: "))
    lista_valores.append(valor)

    porcentagem = int(input("Digite o valor da porcentagem de juros: "))
    lista_porcentagem.append(porcentagem)

    intervalo_pagamentos = int(
        input("Digite a cada quanto tempo o cliente irá pagar (em dias): ")
    )
    lista_pagamento.append(intervalo_pagamentos)

    parcelas = int(input("Digite o número de parcelas: "))
    lista_parcelas.append(parcelas)

    valor_total = valor
    valor_parcela = (valor_total / parcelas) + (porcentagem / 100 * valor)
    formatted_valor_parcela = round(valor_parcela, 2)
    lista_valor_parcelas.append(formatted_valor_parcela)

    total_a_pagar = valor_total * (1 + porcentagem / 100)
    lista_valor_total.append(total_a_pagar)

    data_hoje = datetime.now()
    data_hoje_form1 = datetime.strftime(data_hoje, "%d/%m/%y")
    data_hoje_form2 = datetime.strptime(data_hoje_form1, "%d/%m/%y")
    lista_data_cadastro.append(data_hoje_form2)

    # Retorne os valores relevantes
    return (
        cliente,
        nmr_cliente,
        valor,
        porcentagem,
        intervalo_pagamentos,
        parcelas,
        formatted_valor_parcela,
        total_a_pagar,
        data_hoje_form2,
    )



def adicionar_dados_planilha_local():
    nome_arquivo = "clientes.xlsx"

    try:
        workbook = openpyxl.load_workbook(nome_arquivo)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active
    def find_next_empty_row(sheet):
        for row in range(1, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value is None:
                return row
        return sheet.max_row + 1

    # Encontrar a próxima linha vazia
    proxima_linha = find_next_empty_row(sheet)

    # Adicione os dados dos clientes à planilha a partir da próxima linha vazia
    for i in range(len(lista_clientes)):
        sheet.cell(row=proxima_linha + i, column=1, value=lista_clientes[i])
        sheet.cell(row=proxima_linha + i, column=2, value=lista_nmr_clientes[i])
        sheet.cell(row=proxima_linha + i, column=3, value=lista_valores[i])
        sheet.cell(row=proxima_linha + i, column=4, value=lista_porcentagem[i])
        sheet.cell(row=proxima_linha + i, column=5, value=lista_pagamento[i])
        sheet.cell(row=proxima_linha + i, column=6, value=lista_parcelas[i])
        sheet.cell(row=proxima_linha + i, column=7, value=lista_valor_parcelas[i])
        sheet.cell(row=proxima_linha + i, column=8, value=lista_valor_total[i])
        sheet.cell(row=proxima_linha + i, column=9, value=lista_data_cadastro[i])

    # Salve a planilha atualizada na pasta do projeto
    workbook.save(nome_arquivo)
    workbook.close()


def identificar_parcelas():
    nome_arquivo = "clientes.xlsx"

    try:
        workbook = openpyxl.load_workbook(nome_arquivo)
    except FileNotFoundError:
        print("A planilha 'clientes.xlsx' não foi encontrada.")
        exit(1)

    sheet = workbook.active

    if sheet.max_row < 2:
        print("A planilha não possui dados suficientes.")
        exit(1)

    data_hoje = datetime.now()
    data_hoje_form1 = datetime.strftime(data_hoje, "%d/%m/%y")
    data_hoje_form2 = datetime.strptime(data_hoje_form1, "%d/%m/%y")

    clientes_match = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[8] is None or row[4] is None or row[4] == 0 or row[5] is None or row[6] is None:
            continue

        nome_cliente = row[0]
        celular_cliente = row[1]
        valor_parcela = row[6]
        intervalo_pagamento = row[4]
        total_parcelas = row[5]

        # Tentativa de converter a data de cadastro
        data_cadastro_str = row[8]



        dias_desde_cadastro = (data_hoje_form2 - data_cadastro_str).days
        total_dias_para_pagar = intervalo_pagamento * total_parcelas

        if 0 < dias_desde_cadastro <= total_dias_para_pagar and dias_desde_cadastro % intervalo_pagamento == 0:
            numero_parcela = dias_desde_cadastro // intervalo_pagamento
            clientes_match.append({
                'Nome': nome_cliente,
                'Celular': celular_cliente,
                'Parcela': numero_parcela,
                'Valor da Parcela': valor_parcela
            })

    workbook.close()
    return clientes_match

def enviar_sms_cliente(nome_cliente, numero_celular, numero_parcela, valor_parcela):
    account_sid = "ACae6b3430341fde63009bb4ccb9881310"
    auth_token = "0c517bd637320f5e88532f7ad523f3b9"
    twilio_phone_number = "+14783304454"
    client_twilio = Client(account_sid, auth_token)

    mensagem = f"Olá {nome_cliente}, hoje expira a data de pagamento da sua {numero_parcela}° parcela no valor de R${valor_parcela}."

    message = client_twilio.messages.create(
        body=mensagem,
        from_=twilio_phone_number,
        to=f'+{numero_celular}'
    )
    print(f"Mensagem enviada para {nome_cliente} ({numero_celular}): {message.sid}")

