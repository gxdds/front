import tkinter as tk
from functions import coletar_dados_cliente, enviar_sms_cliente, identificar_parcelas
import openpyxl
import tkinter as tk
from datetime import datetime, timedelta

def abrir_janela_addcliente():
    janela_add = tk.Toplevel()
    janela_add.geometry("800x490")
    janela_add.configure(bg='#88A3A4')
    janela_add.title('Adicionar Cliente')
    janela_add.rowconfigure([0, 1, 2, 3, 4, 5, 6], weight=1)
    janela_add.columnconfigure([0, 1, 2, 3, 4, 5, 6], weight=1)

    def add_button():
        try:
            nome = input_nome.get()
            numero = int(input_numero.get())
            valoremprestado = int(input_valoremprestado.get())
            porcentagem = int(input_porcentagem.get())
            dias = int(input_dias.get())
            parcelas = int(input_parcelas.get())
        except ValueError:
            # Se ocorrer um erro de valor inválido, exiba uma mensagem de erro na interface
            janela_mensagem_erro = tk.Toplevel()
            janela_mensagem_erro.geometry("500x30")
            janela_mensagem_erro.title("Erro")
            label_erro = tk.Label(janela_mensagem_erro,
                                  text="Confira os campos - Além do campo NOME todos devem ser preenchidos com números.")
            label_erro.grid(row=0, column=1, sticky="NSEW")
            return
        lista_clientes = []
        lista_nmr_clientes = []
        lista_valores = []
        lista_porcentagem = []
        lista_pagamento = []
        lista_parcelas = []
        lista_valor_parcelas = []
        lista_valor_total = []
        lista_data_cadastro = []
        lista_clientes.append(nome)
        lista_nmr_clientes.append(numero)
        lista_valores.append(valoremprestado)
        lista_porcentagem.append(porcentagem)
        lista_pagamento.append(dias)
        lista_parcelas.append(parcelas)
        valor_total = valoremprestado
        valor_parcela = (valor_total / parcelas) + (porcentagem / 100 * valoremprestado)
        formatted_valor_parcela = round(valor_parcela, 2)
        lista_valor_parcelas.append(formatted_valor_parcela)
        total_a_pagar = valor_total * (1 + porcentagem / 100)
        lista_valor_total.append(total_a_pagar)
        data_hoje = datetime.now()
        data_hoje_form1 = datetime.strftime(data_hoje, "%d/%m/%y")
        data_hoje_form2 = datetime.strptime(data_hoje_form1, "%d/%m/%y")
        lista_data_cadastro.append(data_hoje_form1)

        nome_arquivo = "clientes.xlsx"

        try:
            workbook = openpyxl.load_workbook(nome_arquivo)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        sheet = workbook.active

        def find_next_empty_row(sheet):
            for row in range(1, sheet.max_row + 1):
                if sheet.cell(row=row, column=1).value is None:
                    return row
            return sheet.max_row + 1

        # Encontrar a próxima linha vazia
        proxima_linha = find_next_empty_row(sheet)

        # Adicione os dados dos clientes à planilha a partir da próxima linha vazia
        for i in range(len(lista_clientes)):
            sheet.cell(row=proxima_linha + i, column=1, value=lista_clientes[i])
            sheet.cell(row=proxima_linha + i, column=2, value=lista_nmr_clientes[i])
            sheet.cell(row=proxima_linha + i, column=3, value=lista_valores[i])
            sheet.cell(row=proxima_linha + i, column=4, value=lista_porcentagem[i])
            sheet.cell(row=proxima_linha + i, column=5, value=lista_pagamento[i])
            sheet.cell(row=proxima_linha + i, column=6, value=lista_parcelas[i])
            sheet.cell(row=proxima_linha + i, column=7, value=lista_valor_parcelas[i])
            sheet.cell(row=proxima_linha + i, column=8, value=lista_valor_total[i])
            sheet.cell(row=proxima_linha + i, column=9, value=lista_data_cadastro[i])
        try:
            # Salve a planilha atualizada na pasta do projeto
            workbook.save(nome_arquivo)
            workbook.close()

            janela_mensagem = tk.Toplevel()
            janela_mensagem.geometry("395x80")
            janela_mensagem.rowconfigure(0, weight=1)
            janela_mensagem.columnconfigure(0, weight=1)
            janela_mensagem.title("Cliente adicionado com sucesso!")
            label_sucesso = tk.Label(janela_mensagem,
                                     text=f"O cliente foi adicionado com sucesso!")
            label_sucesso.grid(row=0, column=0, sticky="NSEW")

        except Exception as e:
            janela_mensagem = tk.Toplevel()
            janela_mensagem.geometry("500x90")
            janela_mensagem.rowconfigure(0, weight=1)
            janela_mensagem.columnconfigure(0, weight=1)
            janela_mensagem.title("Erro")
            label_erro1 = tk.Label(janela_mensagem,
                                     text=f"Confira os campos - O único que aceita letras é o campo NOME. \n Não se esqueça de estar com a planilha fechada enquanto adiciona o novo usuário \n para que ele possa ser salvo com sucesso. ")
            label_erro1.grid(row=0, column=0, sticky="NSEW")
            label_erro2 = tk.Label(janela_mensagem, text=f"Código de erro: {str(e)}")
            label_erro2.grid(row=1, column=0, sticky="NSEW")

        return input_nome.get(), input_numero.get(), input_valoremprestado.get(), input_porcentagem.get(), input_dias.get(), input_parcelas.get()

    titulo_fonte = ("bold", 12)
    label_nome = tk.Label(janela_add, text="Digite o nome do cliente: ", font=titulo_fonte, bg='#88A3A4')
    input_nome = tk.Entry(janela_add)

    label_numero = tk.Label(janela_add, text="Digite o número do cliente (55+DDD+CEL): ", font=titulo_fonte, bg='#88A3A4')
    input_numero = tk.Entry(janela_add)

    label_valoremprestado = tk.Label(janela_add, text="Digite o valor emprestado ao cliente: ", font=titulo_fonte, bg='#88A3A4')
    input_valoremprestado = tk.Entry(janela_add)

    label_porcentagem = tk.Label(janela_add, text="Digite a porcentagem de juros: ", font=titulo_fonte, bg='#88A3A4')
    input_porcentagem = tk.Entry(janela_add)

    label_dias = tk.Label(janela_add, text="Digite a cada quanto tempo o cliente irá pagar (em dias): ", font=titulo_fonte, bg='#88A3A4')
    input_dias = tk.Entry(janela_add)

    label_parcelas = tk.Label(janela_add, text="Digite quantas parcelas serão: ", font=titulo_fonte, bg='#88A3A4')
    input_parcelas = tk.Entry(janela_add)

    botao_voltar = tk.Button(janela_add, text="Cancelar", command=janela_add.destroy, width=15, height=2, bg="#A46161")
    botao_adicionar = tk.Button(janela_add, text="Adicionar", command=add_button, width=15, height=2, bg="#94CB91")

    #grids
    label_nome.grid(row=0, column=0)
    label_numero.grid(row=1, column=0)
    label_valoremprestado.grid(row=2, column=0)
    label_porcentagem.grid(row=3, column=0)
    label_dias.grid(row=4, column=0)
    label_parcelas.grid(row=5, column=0)

    input_nome.grid(row=0, column=1, sticky="WE")
    input_numero.grid(row=1, column=1, sticky="WE")
    input_valoremprestado.grid(row=2, column=1, sticky="WE")
    input_porcentagem.grid(row=3, column=1, sticky="WE")
    input_dias.grid(row=4, column=1, sticky="WE")
    input_parcelas.grid(row=5, column=1, sticky="WE")

    botao_voltar.grid(row=3, column=3, sticky="NSWE", padx=30, pady=5)
    botao_adicionar.grid(row=2, column=3, sticky="NSWE", padx=30, pady=5)

def ver_enviar_sms_hj():

    pass

def abrir_planilha():
    pass

janela = tk.Tk()
janela.geometry("448x672")
janela.title("SMSToday")
janela.configure(bg='#88A3A4')


titulo_fonte = ("bold", 18)
titulo_label = tk.Label(janela, text="Bem-vindo ao SMSToday", font=titulo_fonte, bg='#88A3A4')
titulo_label.grid(row=1, column=2, padx=85, pady=20)
titulo_label.rowconfigure(1, weight=1)

botao_adicionarcliente = tk.Button(janela, text="Adicionar novo cliente", command=abrir_janela_addcliente, width= 50, height= 5, bg='lightblue')
botao_adicionarcliente.grid(row=8, column=2, pady=50)

botao_ver_sms = tk.Button(janela, text="Ver/Enviar os SMS de hoje", command=ver_enviar_sms_hj, width= 50, height= 5, bg='lightblue')
botao_ver_sms.grid(row=10, column=2, pady=50)

botao_ver_planilha = tk.Button(janela, text="Ver/Editar a planilha de clientes", command=abrir_planilha, width= 50, height= 5, bg='lightblue')
botao_ver_planilha.grid(row=12, column=2, pady=50)



janela.mainloop()







